# -*- coding: utf-8 -*-
"""导出 SOP 为 Excel：首表对齐《AI算法评分标准_动力电池电芯性能核心参数测量》11 列模版。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT = Path(__file__).resolve().parent / "SOP_动力电池电芯性能核心参数测量.xlsx"

HEADERS = [
    "任务模块",
    "任务名称",
    "任务描述",
    "分值",
    "扣分依据",
    "展现形式（物品+操作）",
    "步骤操作否定项",
    "标准话术（口述）",
    "关键词（必说词汇标红）",
    "任务类型",
    "算法实现方式",
]

DESC_INNER = (
    "将红色夹子夹住1号电芯正极；\n"
    "将黑色夹子夹住1号电芯负极；\n"
    "读取1号电芯内阻值并记录；\n"
    "将红色夹子夹住2号电芯正极；\n"
    "将黑色夹子夹住2号电芯负极；\n"
    "读取2号电芯内阻值并记录；\n"
    "将红色夹子夹住3号电芯正极；\n"
    "将黑色夹子夹住3号电芯负极；\n"
    "读取3号电芯内阻值并记录；\n"
    "将红色夹子夹住4号电芯正极；\n"
    "将黑色夹子夹住4号电芯负极；\n"
    "读取4号电芯内阻值并记录；\n"
    "将红色夹子夹住5号电芯正极；\n"
    "将黑色夹子夹住5号电芯负极；\n"
    "读取5号电芯内阻值并记录；\n"
    "将红色夹子夹住6号电芯正极；\n"
    "将黑色夹子夹住6号电芯负极；\n"
    "读取6号电芯内阻值并记录；"
)

DESC_VOLT = (
    "将红色表笔接1号电芯正极；\n"
    "将黑色表笔接1号电芯负极；\n"
    "读取1号电芯电压值并记录；\n"
    "将红色表笔接2号电芯正极；\n"
    "将黑色表笔接2号电芯负极；\n"
    "读取2号电芯电压值并记录；\n"
    "将红色表笔接3号电芯正极；\n"
    "将黑色表笔接3号电芯负极；\n"
    "读取3号电芯电压值并记录；\n"
    "将红色表笔接4号电芯正极；\n"
    "将黑色表笔接4号电芯负极；\n"
    "读取4号电芯电压值并记录；\n"
    "将红色表笔接5号电芯正极；\n"
    "将黑色表笔接5号电芯负极；\n"
    "读取5号电芯电压值并记录；\n"
    "将红色表笔接6号电芯正极；\n"
    "将黑色表笔接6号电芯负极；\n"
    "读取6号电芯电压值并记录；"
)

NEG_INNER = (
    "未将红色夹子夹2、4、6号电芯的正极；\n"
    "未将黑色夹子夹2、4、6号电芯的负极；"
)

NEG_VOLT = (
    "未将红色表笔接2、4、6号电芯的正极；\n"
    "未将黑色表笔接2、4、6号电芯的负极；"
)

DEDUCT_INNER_MAIN = (
    "【与本条任务描述对应的达标要求】\n"
    "每一只电芯（1～6号）均须完成：红夹接正极、黑夹接负极、读数稳定后记录；顺序须为1→2→3→4→5→6，禁止跳号漏测。\n"
    "\n"
    "【与模版一致的否定项口径】\n"
    + NEG_INNER
    + "\n\n"
    "【扣分累计细则】\n"
    "1）以「步骤操作否定项」及现场考评规则为准，每查实1处扣2分；\n"
    "2）本模块（六个电芯内阻测量）合计最高扣18分，不累加超出部分；\n"
    "3）与上述否定项等价的极性反接、应接未接、夹持明显不到位导致无法有效读数等，按考评认定计项；\n"
    "4）除否定项单列外，若存在漏测、未记录、顺序严重错乱等，由监考按实训细则另行计扣（若有）。"
)

DEDUCT_VOLT_MAIN = (
    "【与本条任务描述对应的达标要求】\n"
    "每一只电芯（1～6号）均须完成：红表笔接正极、黑表笔接负极、万用表置于直流电压合适量程、读数稳定后记录；顺序须为1→2→3→4→5→6。\n"
    "\n"
    "【与模版一致的否定项口径】\n"
    + NEG_VOLT
    + "\n\n"
    "【扣分累计细则】\n"
    "1）每查实1处扣2分；\n"
    "2）本模块（六个电芯电压测量）合计最高扣18分；\n"
    "3）与否定项等价的极性反接、表笔未有效接触等，按考评认定计项。"
)

DEDUCT_SUB_18 = (
    "【模版原文摘要】1个2分，合计扣除18分。\n"
    "【执行说明】与本模块主行「扣分依据」一致：按项每项2分、本模块封顶18分。"
)


def style_header(ws, ncol: int, row: int = 1):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF", size=10)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_body_style(ws, start_row: int, ncol: int):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")


def sheet_column_widths(ws, widths: list[float | int]):
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = float(w)


def build_template_rows():
    """与参考 xlsx 行结构对应：主任务行 + 扣分细则补充行（内阻/电压各一行）。"""
    rows = []

    def row(
        m,
        name,
        desc,
        score,
        deduct,
        show,
        neg,
        speech,
        kw,
        kind,
        algo,
    ):
        return [m, name, desc, score, deduct, show, neg, speech, kw, kind, algo]

    # —— 准备阶段 ——
    rows.append(
        row(
            "准备阶段\n(10分)",
            "安全防护",
            "佩戴耐磨手套。",
            2,
            "扣分细则：未佩戴或未规范佩戴耐磨手套，本项2分全扣（该项不得分）。\n"
            "补充说明：手套应完好；佩戴到位后方可接触万用表、内阻仪及电芯。",
            "耐磨手套、学员双手、工位台面",
            "开始操作设备/电芯前未佩戴手套；佩戴后仍裸手接触带电测量端（经认定）。",
            "（示例，以教师发布为准）口述：已检查手套完好，现在佩戴。",
            "手套、佩戴、防护",
            "实操",
            "视觉：手套穿戴状态；可选手部与设备交互区域检测",
        )
    )
    rows.append(
        row(
            "",
            "连接万用表线束",
            "1、红色线束连接最右边插孔；\n2、黑色线束连接右二插孔；",
            2,
            "扣分细则：红色未插「最右侧」孔位，或黑色未插「右数第二」孔位，或未插紧导致松动，本项2分全扣。\n"
            "说明：以本工位万用表丝印为准，插接到位后轻拉确认不脱落。",
            "万用表、红/黑表笔插头、插孔面板",
            "插孔与颜色规定不符；插头虚接经提示仍未纠正。",
            "口述红、黑线分别对应的插孔位置（以标准话术为准）。",
            "红色、最右边、黑色、右二、插紧",
            "实操",
            "插孔类别/颜色识别 + 插头插入深度或锁定状态估计",
        )
    )
    rows.append(
        row(
            "",
            "万用表校零",
            "打到蜂鸣挡，红黑表笔对接，万用表发出蜂鸣声；",
            3,
            "扣分细则：未旋至蜂鸣挡完成短接校核；或短接后无声且未排查档位/表笔/线束即进入后续步骤，本项3分全扣。",
            "万用表旋钮、蜂鸣挡、红/黑表笔金属笔尖",
            "未在蜂鸣挡对接；无声仍继续测量。",
            "口述校零结果：蜂鸣正常/已排故（以标准话术为准）。",
            "蜂鸣挡、对接、蜂鸣声",
            "实操",
            "旋钮档位识别 + 笔尖距离/接触 + 蜂鸣事件检测",
        )
    )
    rows.append(
        row(
            "",
            "电池内阻测试仪线束连接",
            "线束插入内阻测试仪，拧紧；",
            3,
            "扣分细则：线束未完全插入指定接口，或未按规程拧紧防松，本项3分全扣。\n"
            "说明：松动易导致读数跳变或误判为已完成测量。",
            "电池内阻测试仪、测试线插头/锁紧螺母",
            "未插入、未拧紧或反插（以仪器标识为准）。",
            "口述「已插入并拧紧」（以标准话术为准）。",
            "内阻测试仪、插入、拧紧",
            "实操",
            "接口区域 ROI + 旋转拧紧动作检测",
        )
    )

    # —— 内阻 主行 + 扣分补充行 ——
    rows.append(
        row(
            "测量电池内阻\n(36分)",
            "六个电芯内阻测量",
            DESC_INNER,
            36,
            DEDUCT_INNER_MAIN,
            "内阻测试仪、红/黑鳄鱼夹、1～6号电芯、记录表；逐号夹持—读数—记录",
            NEG_INNER,
            "逐号报读或确认测量完成（以教师发布标准话术为准）。",
            "红色夹子、正极、黑色夹子、负极、内阻、记录",
            "实操",
            "夹具颜色与极性 + 电芯编号区域 + 读数停留 + 记录动作（算法参数由系统配置）",
        )
    )
    rows.append(
        row(
            "",
            "",
            "",
            "",
            DEDUCT_SUB_18,
            "",
            "",
            "",
            "",
            "",
            "同上（封顶18分项）",
        )
    )

    # —— 电压 主行 + 扣分补充行 ——
    rows.append(
        row(
            "测量电池电压\n(36分)",
            "六个电芯电压测量",
            DESC_VOLT,
            36,
            DEDUCT_VOLT_MAIN,
            "万用表直流电压档、红/黑表笔、1～6号电芯、记录表；逐号接触—读数—记录",
            NEG_VOLT,
            "逐号报读或确认测量完成（以教师发布标准话术为准）。",
            "红色表笔、正极、黑色表笔、负极、电压、记录",
            "实操",
            "表笔颜色与极性 + 电芯编号 + 屏幕读数稳定帧 + 记录动作",
        )
    )
    rows.append(
        row(
            "",
            "",
            "",
            "",
            DEDUCT_SUB_18,
            "",
            "",
            "",
            "",
            "",
            "同上（封顶18分项）",
        )
    )

    # —— 工位恢复（模版写「工位回复」）——
    rows.append(
        row(
            "工位回复\n(12分)",
            "恢复操作台（万用表关机）",
            "万用表旋钮打到OFF挡位。",
            4,
            "扣分细则：测量结束后未将万用表旋钮旋至 OFF（或等效关机挡），本小项4分全扣。",
            "万用表旋钮、OFF标识",
            "旋钮仍停留在电压/蜂鸣等测量挡即离开工位或进入下一考核单元。",
            "口述已关机/OFF（以标准话术为准）。",
            "OFF、关机",
            "实操",
            "旋钮角度/OFF 区域分类",
        )
    )
    rows.append(
        row(
            "",
            "拔出万用表线束",
            "自万用表端规范拔出红、黑表笔插头，并放置于指定收纳位。",
            4,
            "扣分细则：未拔出万用表表笔线束或拔出不完整，本小项4分全扣。",
            "万用表、表笔插头",
            "线束仍插在万用表上即离场。",
            "口述线束已拆除（以标准话术为准）。",
            "拔出、万用表线束",
            "实操",
            "插头在位检测",
        )
    )
    rows.append(
        row(
            "",
            "拔出电池内阻测试仪线束",
            "自内阻测试仪端规范拔出测试线束，并整理放置。",
            4,
            "扣分细则：未拔出内阻测试仪测试线束，本小项4分全扣。",
            "内阻测试仪、测试线插头",
            "测试线仍插在仪器上即离场。",
            "口述内阻仪线束已拆除（以标准话术为准）。",
            "拔出、内阻测试仪",
            "实操",
            "插头在位检测",
        )
    )

    # —— 收尾 ——
    rows.append(
        row(
            "收尾\n(6分)",
            "清理整理工作台面",
            "清理杂物与碎屑；表笔/夹具/线材归位；台面整洁、无工具遗落，符合实训室6S要求。",
            6,
            "扣分细则：工位凌乱、工具或线材未归位、存在安全隐患（如金属导体裸露乱放），本项6分按细则扣减至不得分（以监考评定为准）。",
            "工位台面、工具托盘、线材收纳",
            "台面未整理即结束考核；关键器具未归位。",
            "口述现场已整理完毕（以标准话术为准）。",
            "清理、整理、归位",
            "实操",
            "台面整洁度/物品类别与位置估计（算法阈值由系统配置）",
        )
    )

    return rows


def main():
    wb = Workbook()

    # --- Sheet1：与模版一致的 11 列 ---
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(HEADERS)
    style_header(ws1, len(HEADERS))
    for r in build_template_rows():
        ws1.append(r)
    apply_body_style(ws1, 2, len(HEADERS))
    ws1.freeze_panes = "A2"
    sheet_column_widths(
        ws1,
        [14, 16, 52, 8, 44, 22, 28, 18, 18, 8, 26],
    )

    # --- 文档说明 ---
    ws0 = wb.create_sheet("文档说明", 1)
    info = [
        ["SOP 名称", "动力电池电芯性能核心参数测量"],
        ["Excel 说明", "Sheet1 列结构与《AI算法评分标准_动力电池电芯性能核心参数测量》一致；已补全扣分依据、展现形式、否定项及算法列供实施参考。"],
        ["总分", "100 分"],
        ["", ""],
        ["1. 目的", "规范电芯内阻、电压测量与工位恢复流程，与考核评分项一一对应。"],
        ["2. 适用范围", "动力电池电芯性能核心参数测量实训与考核。"],
    ]
    for i, row in enumerate(info, start=1):
        ws0.cell(row=i, column=1, value=row[0])
        ws0.cell(row=i, column=2, value=row[1])
        ws0.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws0.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 14
    ws0.column_dimensions["B"].width = 80

    # --- 测量记录表 ---
    ws4 = wb.create_sheet("测量记录表", 2)
    ws4.append(["电芯编号", "内阻值", "电压值", "测量时间", "测量人", "备注"])
    style_header(ws4, 6)
    for i in range(1, 7):
        ws4.append([f"{i}号", "", "", "", "", ""])
    for r in range(1, ws4.max_row + 1):
        for c in range(1, 7):
            ws4.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 18
    ws4.column_dimensions["E"].width = 12
    ws4.column_dimensions["F"].width = 24

    wb.save(OUT)
    print("ok", OUT.stat().st_size)


if __name__ == "__main__":
    main()
